"""Parser CSV/TSV/XLSX per sample sheet (PRD FR-005).

Il sample sheet contiene la provenance essenziale del disegno: e per questo che
N-Truth legge tabelle e non solo testo (PRD 11.2). Le formule vengono disinnescate
prima di entrare nel Document IR (PRD NFR-13).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from pathlib import Path

from ntruth.ingest.safety import neutralize_formula
from ntruth.parsers.base import ParseFailure, RawDocument, RawTable
from ntruth.schemas.document import ParserStatus

MAX_ROWS = 200_000
MAX_COLUMNS = 512


class CsvParser:
    name = "csv"

    def supports(self, path: Path, media_type: str) -> bool:
        return path.suffix.lower() in {".csv", ".tsv"}

    def parse(self, path: Path) -> RawDocument:
        doc = RawDocument(parser=self.name)
        try:
            content = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            content = path.read_text(encoding="latin-1")
            doc.warnings.append("codifica non UTF-8: letto come latin-1")
        except OSError as exc:  # pragma: no cover
            raise ParseFailure(path, f"lettura fallita ({exc})") from exc

        delimiter = "\t" if path.suffix.lower() == ".tsv" else _sniff_delimiter(content)
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        table = _build_table(path.stem, reader, doc)
        if table is not None:
            doc.tables.append(table)
        else:
            doc.status = ParserStatus.FAILED
            doc.warnings.append("file tabellare vuoto")
        return doc


class XlsxParser:
    name = "xlsx"

    def supports(self, path: Path, media_type: str) -> bool:
        return path.suffix.lower() == ".xlsx"

    def parse(self, path: Path) -> RawDocument:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dipendenza dichiarata
            raise ParseFailure(path, "openpyxl non installato") from exc

        doc = RawDocument(parser=self.name)
        try:
            # data_only=False conserva la formula come testo inerte. openpyxl non
            # la esegue; _build_table la neutralizza prima del Document IR.
            # read_only limita la memoria (NFR-07/NFR-13).
            workbook = load_workbook(str(path), data_only=False, read_only=True)
        except Exception as exc:
            raise ParseFailure(path, f"XLSX illeggibile ({type(exc).__name__})") from exc

        try:
            for sheet in workbook.worksheets:
                rows = (
                    ["" if cell is None else str(cell) for cell in row]
                    for row in sheet.iter_rows(values_only=True)
                    if any(value is not None and str(value).strip() for value in row)
                )
                table = _build_table(sheet.title, rows, doc, sheet=sheet.title)
                if table is None:
                    doc.warnings.append(f"foglio '{sheet.title}' vuoto")
                    continue
                doc.tables.append(table)
        finally:
            workbook.close()

        if not doc.tables:
            doc.status = ParserStatus.FAILED
            doc.warnings.append("nessun foglio leggibile")
        return doc


def _sniff_delimiter(content: str) -> str:
    sample = content[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def _build_table(
    name: str, rows: Iterable[list[str]], doc: RawDocument, sheet: str | None = None
) -> RawTable | None:
    iterator = iter(rows)
    try:
        header_raw = next(iterator)
    except StopIteration:
        return None
    if len(header_raw) > MAX_COLUMNS:
        doc.warnings.append(f"{name}: {len(header_raw)} colonne oltre il limite, troncate")
        header_raw = header_raw[:MAX_COLUMNS]
    formula_cells = 0
    safe_header: list[str] = []
    for value in header_raw:
        safe, changed = neutralize_formula(str(value))
        formula_cells += int(changed)
        safe_header.append(safe.strip())
    header = _unique_headers(safe_header)
    if not header:
        return None

    table = RawTable(name=name, sheet=sheet, columns=header)
    for index, raw in enumerate(iterator):
        if index >= MAX_ROWS:
            table.warnings.append(f"righe oltre {MAX_ROWS} ignorate")
            doc.status = ParserStatus.PARTIAL
            break
        values = list(raw[: len(header)])
        values.extend([""] * (len(header) - len(values)))
        record: dict[str, str] = {}
        for column, value in zip(header, values, strict=True):
            safe, changed = neutralize_formula(str(value))
            formula_cells += int(changed)
            record[column] = safe.strip()
        table.rows.append(record)

    if formula_cells:
        message = f"{name}: {formula_cells} celle con prefisso di formula disinnescate"
        table.warnings.append(message)
        doc.warnings.append(message)
    return table


def _unique_headers(raw: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for index, value in enumerate(raw):
        name = (value or "").strip() or f"col_{index + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out
